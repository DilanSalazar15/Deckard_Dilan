import random
import pandas as pd
from datetime import date, timedelta
import requests
import icalendar
import re
from openpyxl import load_workbook

ICAL_FEED_URL = "https://deckard.bamboohr.com/feeds/feed.php?id=9be94f136c71a44452c47eae8acd4903"

# Diccionario de nombres normalizados
normalized_names = {
    "juliana valencia rodriguez": "Juliana Valencia",
    "carlos montenegro galan": "Carlos Montenegro",
    "sara osorio castaño": "Sara Osorio",
    "juan pineda llano": "Juan Pineda",
    "alejandro gámez pérez": "Enrique Gamez",
    "hugo pulgarín lópez": "Hugo Pulgarin",
    "juan amud vásquez": "Juan Amud",
    "juan velásquez": "Juan Velasquez",
    "laura posada bolívar": "Laura Posada",
    "manuel Valencia restrepo": "Manuel Valencia",
    "santiago correa velásquez": "Santiago Correa",
    "santiago gómez restrepo": "Santiago Gomez",
    "nicolás arango alzate": "Nicolas Arango",
    "alci monterrosa meza": "Alcibiades Monterrosa",
    "alejandro pineda gil": "Alejandro Pineda",
    "alejo vélez londoño": "Alejandro Velez",
    "juan ruiz pérez": "Juan David Ruiz",
    "luisa ruiz taborda": "Luisa Ruiz",
    "manu durán cuadros": "Manuela Duran",
    "valeria vásquez hernández": "Valeria Vasquez"
}

def normalize_name(name):
    return name.lower().strip()

def get_absent_people_from_ical():
    try:
        response = requests.get(ICAL_FEED_URL)
        response.raise_for_status()

        calendar = icalendar.Calendar.from_ical(response.text)
        absent_people = {}
        pattern = re.compile(r"^(.*?)\s*\((.*?)\s*-\s*\d+\s*days?\)")

        for event in calendar.walk('VEVENT'):
            summary = str(event.get('summary', ''))
            start_date = event.get('dtstart').dt
            end_date = event.get('dtend')

            if end_date:
                end_date = end_date.dt - timedelta(days=1)
            else:
                end_date = start_date

            match = pattern.match(summary)
            if match:
                raw_name = match.group(1).strip()
                normalized_name = normalize_name(raw_name)

                if normalized_name in normalized_names:
                    corrected_name = normalized_names[normalized_name]
                    for single_date in pd.date_range(start_date, end_date):
                        date_str = single_date.strftime('%Y-%m-%d')
                        if date_str not in absent_people:
                            absent_people[date_str] = set()
                        absent_people[date_str].add(corrected_name)

        return absent_people
    except requests.exceptions.RequestException as e:
        print(f"Error al descargar iCalendar: {e}")
        return {}

# Leer planificación con pandas
planning = pd.read_excel('Address Mapping Work Planning.xlsx', sheet_name='Planning_List')

# Leer comentarios con openpyxl
wb = load_workbook("Address Mapping Work Planning.xlsx")
ws = wb["Planning_List"]

comentarios = []
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):  # Columna D
    cell = row[0]
    comentarios.append(cell.comment.text if cell.comment else "")

planning["Comment"] = comentarios

# Obtener ausencias
absent_people = get_absent_people_from_ical()

# Leer otras hojas
work_allocation = pd.read_excel('Work_allocation_Address_Mapping_Code.xlsx', 'Allocation')
status_report = pd.read_excel('Work_allocation_Address_Mapping_Code.xlsx', 'Lists')
maintenance_list = pd.read_excel('Work_allocation_Address_Mapping_Code.xlsx', 'Maintenance_List')

work_allocation['Date'] = pd.to_datetime(work_allocation['Date'], errors='coerce')

# Crear listas
analistas = planning['Analysts'].dropna().unique().tolist()
letterAnalysts = planning['Letters'].dropna().unique().tolist()
propertyAnalysts = planning['PropertyscapeAnalysts'].dropna().unique().tolist()
AustraliaAnalyst = planning['Australia'].dropna().unique().tolist()
qaers = planning['QAers'].dropna().unique().tolist()
BIAnalysts = planning['BI Analysts'].dropna().unique().tolist()
half_storypoint_analysts = planning['New Hires'].dropna().unique().tolist()

today = date.today()
planning['Date'] = pd.to_datetime(planning['Date'])
new_week = planning.loc[planning['Date'].dt.date > today]

# Agregar entradas a Maintenance_List desde comentarios
for index, row in planning.iterrows():
    if row['Project'] == 'Maintenance Several Projects' and pd.notna(row['Comment']):
        fecha = pd.to_datetime(row['Date']).date()
        lineas = row['Comment'].split('\n')
        proyectos = [line.strip() for line in lineas[1:] if line.strip()]
        for proyecto in proyectos:
            nueva_fila = pd.DataFrame({
                'Date': [fecha],
                'Project': [proyecto]
            })
            maintenance_list = pd.concat([maintenance_list, nueva_fila], ignore_index=True)

# Asignar analistas
for index, row in new_week.iterrows():
    proyecto = row['Project']
    num_analistas = round(row['Story Points'])
    fecha = row['Date'].date()
    fecha_str = fecha.strftime('%Y-%m-%d')
    ausentes = absent_people.get(fecha_str, set())

    def filtrar_disponibles(lista):
        return list(set(lista) - set(work_allocation[work_allocation['Date'] == fecha]['Analyst'].tolist()))

    if row['Project'] == 'Leave':
        for ausente in ausentes:
            nueva_fila = pd.DataFrame({
                'Date': [fecha],
                'Week': [fecha.isocalendar()[1]],
                'Project': ['Leave'],
                'Analyst': [ausente],
                'Type': ['']
            })
            work_allocation = pd.concat([work_allocation, nueva_fila], ignore_index=True)
        continue

    if isinstance(proyecto, str) and proyecto.strip() in ['AUS - City of Adelaide', 'AUS - City of Brisbane', 'AUS - City of Marion']:
        disponibles = filtrar_disponibles(AustraliaAnalyst)
    elif proyecto == 'Letter Campaign':
        disponibles = filtrar_disponibles(letterAnalysts)
    elif proyecto == 'Propertyscape':
        disponibles = filtrar_disponibles(propertyAnalysts)
    elif proyecto == 'BI Analysis':
        disponibles = filtrar_disponibles(BIAnalysts)
    elif proyecto == 'Quality Assurance':
        disponibles = filtrar_disponibles(qaers)
    else:
        disponibles = filtrar_disponibles(analistas)

    random.shuffle(disponibles)
    normales = [a for a in disponibles if a not in half_storypoint_analysts]
    medios = [a for a in disponibles if a in half_storypoint_analysts]

    asignados = []
    toggle = True

    while num_analistas > 0:
        if num_analistas == 1:
            if toggle and normales:
                asignado = random.choice(normales)
                normales.remove(asignado)
                asignados.append(asignado)
                num_analistas -= 1
            elif not toggle and len(medios) >= 2:
                asignado1 = medios.pop()
                asignado2 = medios.pop()
                asignados.extend([asignado1, asignado2])
                num_analistas -= 1
        else:
            if normales:
                asignado = random.choice(normales)
                normales.remove(asignado)
                asignados.append(asignado)
                num_analistas -= 1
            elif len(medios) >= 2:
                asignado1 = medios.pop()
                asignado2 = medios.pop()
                asignados.extend([asignado1, asignado2])
                num_analistas -= 1
            else:
                break
        toggle = not toggle

    for analista in asignados:
        nueva_fila = pd.DataFrame({
            'Analyst': [analista],
            'Project': [proyecto],
            'Date': [fecha],
            'Type': [row['Category']],
            'Week': [fecha.isocalendar()[1]],
        })
        work_allocation = pd.concat([work_allocation, nueva_fila], ignore_index=True)

# QA tracking
work_allocation['Date'] = pd.to_datetime(work_allocation['Date'])
month = today.month
qaers_month_work = work_allocation.loc[work_allocation['Date'].dt.month == month]
QA_control = pd.DataFrame()

for qaer in qaers:
    allocated_in_month = qaers_month_work[(qaers_month_work['Project'] == 'Quality Assurance') &
                                          (qaers_month_work['Analyst'] == qaer)].shape[0]
    qaerRegister = pd.DataFrame({
        'QAer': [qaer],
        'Allocated in Month': [allocated_in_month]
    })
    QA_control = pd.concat([QA_control, qaerRegister], ignore_index=True)

work_allocation['Date'] = work_allocation['Date'].dt.date
maintenance_list['Date'] = pd.to_datetime(maintenance_list['Date']).dt.date

# Guardar en Excel
data_dict = {
    "Allocation": work_allocation,
    "Maintenance_List": maintenance_list,
    "Lists": status_report,
    "QAers": QA_control
}

# Asignar instrucciones Realist aleatoriamente cada día
realist_messages = [
    "Realist account: Shartpau - Deckard2024",
    "Realist account: PT812627 - @Deckard2024",
    "Realist account: PT820539 - @Deckard2024"
]

# Crear la columna Instructions vacía
work_allocation["Instructions"] = ""

# Agrupar por fecha
for fecha, grupo in work_allocation.groupby("Date"):
    analistas_disponibles = grupo["Analyst"].dropna().unique().tolist()
    if len(analistas_disponibles) >= 3:
        seleccionados = random.sample(analistas_disponibles, 3)
        mensajes_del_dia = realist_messages.copy()
        random.shuffle(mensajes_del_dia)  # Mezcla aleatoriamente los mensajes
        for analista, mensaje in zip(seleccionados, mensajes_del_dia):
            mask = (work_allocation["Date"] == fecha) & (work_allocation["Analyst"] == analista)
            work_allocation.loc[mask, "Instructions"] = mensaje

writer = pd.ExcelWriter('Work_allocation_Address_Mapping_Code.xlsx', engine='xlsxwriter')
workbook = writer.book
fmt_header = workbook.add_format({
    'bold': True,
    'fg_color': '#002060',
    'font_color': '#FFFFFF',
})

for sheet_name, df in data_dict.items():
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    for col_num, data in enumerate(df.columns.values):
        writer.sheets[sheet_name].write(0, col_num, data, fmt_header)

worksheetAllocation = writer.sheets['Allocation']
worksheetAllocation.freeze_panes(1, 1)
worksheetAllocation.autofit()
worksheetAllocation.autofilter(0, 0, data_dict['Allocation'].shape[0], len(data_dict['Allocation'].columns) - 1)

worksheetMaintenance = writer.sheets['Maintenance_List']
worksheetMaintenance.autofit()
worksheetMaintenance.autofilter(0, 0, data_dict['Maintenance_List'].shape[0], len(data_dict['Maintenance_List'].columns) - 1)

worksheetLists = writer.sheets['Lists']
formato_encabezados1 = workbook.add_format({'bg_color': '#c0504d', 'font_color': '#FFFFFF'})
formato_encabezados2 = workbook.add_format({'bg_color': '#f79646', 'font_color': '#FFFFFF'})
formato_encabezados3 = workbook.add_format({'bg_color': '#4bacc6', 'font_color': '#FFFFFF'})

worksheetLists.conditional_format('A1:D1', {'type': 'no_blanks', 'format': fmt_header})
worksheetLists.conditional_format('E1:E1', {'type': 'no_blanks', 'format': formato_encabezados1})
worksheetLists.conditional_format('F1:F1', {'type': 'no_blanks', 'format': formato_encabezados2})
worksheetLists.conditional_format('G1:G1', {'type': 'no_blanks', 'format': formato_encabezados3})
worksheetLists.autofit()
worksheetLists.autofilter(0, 0, data_dict['Lists'].shape[0], len(data_dict['Lists'].columns) - 1)

worksheetQAers = writer.sheets['QAers']
worksheetQAers.autofit()

writer.close()
