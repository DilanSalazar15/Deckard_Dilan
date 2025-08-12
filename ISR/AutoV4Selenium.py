import pandas as pd
import os
from pathlib import Path
import numpy as np
import openpyxl
from datetime import datetime
import re
import shutil
import time
import subprocess

# --- 1. IMPORTACIÓN DE SELENIUM (CON VERIFICACIÓN) ---
try:
    from selenium import webdriver
    from selenium.webdriver.edge.service import Service as EdgeService
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import WebDriverException, TimeoutException
    from webdriver_manager.microsoft import EdgeChromiumDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# --- 2. NUEVA FUNCIÓN PARA CERRAR EDGE ---
def close_edge_processes():
    """
    Usa el comando taskkill de Windows para forzar el cierre de todos los procesos de Microsoft Edge.
    """
    print("\nIntentando cerrar todos los procesos de Microsoft Edge para asegurar un inicio limpio...")
    try:
        # /F para forzar el cierre, /IM para especificar el nombre de la imagen (proceso)
        # capture_output=True y text=True para evitar que el comando imprima en nuestra consola
        result = subprocess.run(["taskkill", "/F", "/IM", "msedge.exe"], capture_output=True, text=True)
        # El comando tiene éxito (código 0) si cierra procesos.
        # Falla si no encuentra procesos, pero para nosotros eso también es un éxito.
        if result.returncode == 0:
            print("Éxito: Se han cerrado los procesos de Edge que estaban abiertos.")
        elif "no se encontraron tareas" in result.stderr.lower():
            print("Información: Microsoft Edge ya estaba cerrado.")
        else:
            # Imprimir el error si es algo inesperado
            print(f"Aviso: El comando taskkill devolvió un error: {result.stderr}")
        return True
    except FileNotFoundError:
        print("ERROR: El comando 'taskkill' no se encontró. Asegúrate de estar ejecutando en Windows.")
        return False
    except Exception as e:
        print(f"Ocurrió un error inesperado al intentar cerrar Edge: {e}")
        return False

# --- 3. SECCIÓN DE DESCARGA AUTOMÁTICA CON EDGE (MODIFICADA) ---
def download_latest_reports(profile_path: str):
    """Maneja la descarga usando Selenium con un perfil de Edge existente."""
    if not SELENIUM_AVAILABLE:
        print("AVISO: Librerías 'selenium' o 'webdriver-manager' no están instaladas.")
        return False
    
    # --- PASO CLAVE: CERRAR EDGE ANTES DE EMPEZAR ---
    if not close_edge_processes():
        print("No se pudieron cerrar los procesos de Edge. La descarga automática no puede continuar de forma segura.")
        return False
        
    if not profile_path or "PEGA_TU_RUTA" in profile_path:
        print("ERROR: La ruta del perfil de Edge no ha sido configurada.")
        return False

    downloads_path = Path.home() / "Downloads"
    homepage_dest = r"C:\Users\Deckard\OneDrive - deckardtech.com\Documentos\ISR_Automation\ISR Data\Home page"
    project_stats_dest = r"C:\Users\Deckard\OneDrive - deckardtech.com\Documentos\ISR_Automation\ISR Data\Project Stats"

    driver = None
    try:
        print("\n--- Iniciando descarga con Selenium y perfil de Edge ---")
        options = webdriver.EdgeOptions()
        profile_path_obj = Path(profile_path)
        options.add_argument(f"user-data-dir={profile_path_obj.parent}")
        options.add_argument(f"profile-directory={profile_path_obj.name}")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])

        service = EdgeService(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service, options=options)
        
        wait = WebDriverWait(driver, 20)
        print("Perfil de Edge cargado. Procediendo a las descargas...")

        print("\nDescargando CSV de Home page...")
        driver.get("https://cyborg.deckard.com/")
        files_before = set(os.listdir(downloads_path))
        wait.until(EC.element_to_be_clickable((By.ID, "btn_export_task_stats_per_region"))).click()
        print("Botón de CSV presionado. Esperando 10 segundos...")
        time.sleep(10)

        files_after = set(os.listdir(downloads_path))
        new_files = files_after - files_before
        csv_file = next((f for f in new_files if f.startswith("cyborg_job_stats") and f.endswith(".csv")), None)
        if csv_file:
            print(f"Archivo CSV '{csv_file}' descargado. Moviendo...")
            shutil.move(downloads_path / csv_file, os.path.join(homepage_dest, csv_file))
        else: print("AVISO: No se encontró el nuevo archivo CSV descargado.")

        print("\nDescargando Excel de Project Stats...")
        driver.get("https://cyborg.deckard.com/project_stats")
        files_before = set(os.listdir(downloads_path))
        wait.until(EC.element_to_be_clickable((By.ID, "btn_export_project_stats"))).click()
        print("Botón de Excel presionado. Esperando 10 segundos...")
        time.sleep(10)

        files_after = set(os.listdir(downloads_path))
        new_files = files_after - files_before
        xls_file = next((f for f in new_files if f.startswith("cyborg_project_stats") and f.endswith(".xls")), None)
        if xls_file:
            print(f"Archivo Excel '{xls_file}' descargado. Moviendo...")
            shutil.move(downloads_path / xls_file, os.path.join(project_stats_dest, xls_file))
        else: print("AVISO: No se encontró el nuevo archivo Excel descargado.")
            
        driver.quit()
        return True
    except (WebDriverException, TimeoutException) as e:
        print(f"\n--- ERROR DE SELENIUM ---")
        print(f"No se pudo completar la descarga. Causa: {e}")
        if driver: driver.quit()
        return False
    except Exception as e:
        print(f"\n--- ERROR INESPERADO EN SELENIUM ---")
        print(f"Ocurrió un error: {e}")
        if driver: driver.quit()
        return False

# --- 4. SECCIÓN DE PROCESAMIENTO DE DATOS ---
def find_latest_file(folder_path: str):
    try:
        folder = Path(folder_path)
        files = [f for f in folder.glob('*') if f.is_file() and not f.name.startswith('~')]
        if not files: raise FileNotFoundError(f"No se encontraron archivos en: {folder_path}")
        latest_file = max(files, key=os.path.getmtime)
        print(f"Archivo más reciente encontrado en '{folder.name}': {latest_file.name}")
        return str(latest_file)
    except Exception as e:
        print(f"AVISO: No se pudo buscar archivo en '{folder_path}'. Error: {e}")
        return None

def read_source_file(file_path: str):
    if file_path.lower().endswith('.csv'): return pd.read_csv(file_path)
    else: return pd.read_excel(file_path)

def process_project_stats(folder_path: str):
    print("\n--- Procesando Consulta 1: Project Stats ---")
    latest_file_path = find_latest_file(folder_path)
    if not latest_file_path: return None
    try:
        df = read_source_file(latest_file_path)
        if 'rental_type' in df.columns: df = df[df['rental_type'] == 'STR'].copy()
        cols_to_remove = ["state", "county_cousub", "city", "rental_type", "live", "live_left_to_review", "matched", "unique_apns", "marked_as_out_of_scope", "mapped_to_mus", "no_match_found_active", "no_match_found_inactive", "not_live_when_vetting", "predicted_non_residential", "scrape_date_p"]
        df = df.drop(columns=[col for col in cols_to_remove if col in df.columns], errors='ignore')
        if 'phase_1' in df.columns and 'phase_2' in df.columns:
            df['YTRL'] = df['phase_1'] + df['phase_2']
            df = df.drop(columns=['phase_1', 'phase_2'])
        final_order = ["listings_total", "active", "phase_3", "matched_and_active", "active_out_of_scope_no_apn", "YTRL", "place"]
        df = df[[col for col in final_order if col in df.columns]]
        print("--- Consulta 1 completada. ---")
        return df
    except Exception as e:
        print(f"ERROR procesando 'Project Stats': {e}")
        return None

def process_homepage_data(folder_path: str):
    print("\n--- Procesando Consulta 2: Home page ---")
    latest_file_path = find_latest_file(folder_path)
    if not latest_file_path: return None
    try:
        df = read_source_file(latest_file_path)
        cols_to_remove = ["Source.Name", "stats_date", "listings", "resolved", "mapped_properties", "out_of_scope", "may_be_duplicates_of_already_mapped", "may_have_matching_license", "apns_with_unit_numbers_need_review", "remaining_job_stats_ts", "offline_on_review", "no_match_found_3_or_more", "randomly_qa", "completion"]
        df = df.drop(columns=[col for col in cols_to_remove if col in df.columns], errors='ignore')
        required_cols = ['state_p', 'county_p', 'city_p']
        if not all(col in df.columns for col in required_cols): raise KeyError(f"Columnas necesarias faltantes: {[c for c in required_cols if c not in df.columns]}")
        df['state lower case'] = df['state_p'].str.lower()
        df['string name'] = np.where(df['city_p'] == '_', df['state lower case'] + '-' + df['county_p'], df['state lower case'] + '-' + df['county_p'] + '-' + df['city_p'])
        df = df.drop(columns=["state lower case", "state_p", "county_p", "city_p"])
        final_order = ["never_worked_on_yet", "no_match_found_1", "no_match_found_2", "suggest_qa", "outstanding_disputes", "bad_apns", "string name", "may_have_matching_zillow_listings"]
        df = df[[col for col in final_order if col in df.columns]]
        print("--- Consulta 2 completada. ---")
        return df
    except Exception as e:
        print(f"ERROR procesando 'Home page': {e}")
        return None

# --- 5. BLOQUE PRINCIPAL DE EJECUCIÓN ---
if __name__ == "__main__":
    # !! IMPORTANTE !!
    # VERIFICA QUE ESTA RUTA DE TU PERFIL DE EDGE SEA CORRECTA (de edge://version)
    EDGE_PROFILE_PATH = r"C:\Users\Deckard\AppData\Local\Microsoft\Edge\User Data\Default"
    
    download_successful = download_latest_reports(EDGE_PROFILE_PATH)

    if not download_successful:
        print("\n>>> ACCIÓN REQUERIDA (PLAN B) <<<")
        print("La descarga automática falló. Por favor, descarga los dos archivos manualmente.")
        input("Una vez que los archivos estén en tu carpeta de 'Descargas', presiona 'Enter' para continuar...")
        downloads_path = Path.home() / "Downloads"
        homepage_dest = r"C:\Users\Dilan Salazar\Documents\ISR_Automation\ISR Data\Home page"
        project_stats_dest = r"C:\Users\Dilan Salazar\Documents\ISR_Automation\ISR Data\Project Stats"
        latest_csv = max(downloads_path.glob("cyborg_job_stats*.csv"), key=os.path.getctime, default=None)
        if latest_csv: shutil.move(latest_csv, os.path.join(homepage_dest, latest_csv.name))
        latest_xls = max(downloads_path.glob("cyborg_project_stats*.xls"), key=os.path.getctime, default=None)
        if latest_xls: shutil.move(latest_xls, os.path.join(project_stats_dest, latest_xls.name))
    
    print("\n\n--- Iniciando procesamiento de archivos locales ---")
    base_folder = r"C:\Users\Dilan Salazar\Documents\ISR_Automation"
    project_stats_folder = os.path.join(base_folder, "ISR Data", "Project Stats")
    homepage_data_folder = os.path.join(base_folder, "ISR Data", "Home page")
    master_report_path = os.path.join(base_folder, "Identification_Status_Report.xlsx")
    master_path_obj = Path(master_report_path)
    processed_report_path = master_path_obj.parent / f"{master_path_obj.stem}_processed.xlsx"
    try:
        print(f"\nCreando copia de trabajo en: {processed_report_path}")
        shutil.copy2(master_report_path, processed_report_path)
        df_project_stats = process_project_stats(project_stats_folder)
        df_homepage_data = process_homepage_data(homepage_data_folder)
        print("\nActualizando hojas de datos base en el archivo procesado...")
        with pd.ExcelWriter(processed_report_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            if df_project_stats is not None:
                df_project_stats.to_excel(writer, sheet_name='Project Stats', index=False)
            if df_homepage_data is not None:
                df_homepage_data.to_excel(writer, sheet_name='Home page', index=False)
        print("Hojas de datos base actualizadas.")
        print("\nIniciando manipulación de hojas con fecha...")
        wb = openpyxl.load_workbook(processed_report_path)
        date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        dated_sheets = {datetime.strptime(s, '%Y-%m-%d').date(): s for s in wb.sheetnames if date_pattern.match(s) and datetime.strptime(s, '%Y-%m-%d').date() <= datetime.now().date()}
        if not dated_sheets: raise ValueError("No se encontraron hojas con formato AAAA-MM-DD en el archivo.")
        latest_date = max(dated_sheets.keys())
        sheet_to_copy_name = dated_sheets[latest_date]
        today_str = datetime.now().strftime('%Y-%m-%d')
        new_sheet_name = today_str
        if today_str not in wb.sheetnames:
            print(f"Hoja más reciente encontrada: '{sheet_to_copy_name}'")
            sheet_to_copy = wb[sheet_to_copy_name]
            new_sheet = wb.copy_worksheet(sheet_to_copy)
            new_sheet.title = today_str
            sheet_to_copy.sheet_state = 'hidden'
            print(f"Hoja '{sheet_to_copy_name}' duplicada como '{new_sheet_name}' y original ocultada.")
        print(f"\nLimpiando y poblando datos en la hoja '{new_sheet_name}'...")
        ws = wb[new_sheet_name]
        df_ps_lookup = pd.DataFrame(wb['Project Stats'].values); df_hp_lookup = pd.DataFrame(wb['Home page'].values)
        if not df_ps_lookup.empty: df_ps_lookup.columns = df_ps_lookup.iloc[0]; df_ps_lookup = df_ps_lookup[1:].set_index('place')
        if not df_hp_lookup.empty: df_hp_lookup.columns = df_hp_lookup.iloc[0]; df_hp_lookup = df_hp_lookup[1:].set_index('string name')
        for row_idx in range(2, ws.max_row + 1):
            lookup_key = ws[f'AM{row_idx}'].value
            if not lookup_key: continue
            if not df_ps_lookup.empty and lookup_key in df_ps_lookup.index:
                match = df_ps_lookup.loc[lookup_key]
                ws[f'F{row_idx}'], ws[f'G{row_idx}'], ws[f'H{row_idx}'], ws[f'I{row_idx}'], ws[f'J{row_idx}'] = match[['listings_total', 'active', 'phase_3', 'matched_and_active', 'active_out_of_scope_no_apn']].values
                ws[f'P{row_idx}'] = match['YTRL']
            if not df_hp_lookup.empty and lookup_key in df_hp_lookup.index:
                match = df_hp_lookup.loc[lookup_key]
                ws[f'Q{row_idx}'], ws[f'R{row_idx}'], ws[f'S{row_idx}'], ws[f'T{row_idx}'], ws[f'U{row_idx}'], ws[f'V{row_idx}'] = match[['never_worked_on_yet', 'no_match_found_1', 'no_match_found_2', 'suggest_qa', 'outstanding_disputes', 'bad_apns']].values
        print("Datos escritos directamente en la hoja.")
        print("\nLimpiando el archivo para dejar solo las hojas necesarias...")
        sheets_to_keep = ['Project Stats', 'Home page', new_sheet_name]
        for sheet_name in wb.sheetnames[:]:
            if sheet_name not in sheets_to_keep:
                del wb[sheet_name]
                print(f"Hoja '{sheet_name}' eliminada.")
        print("\nGuardando todos los cambios en el archivo final...")
        wb.save(processed_report_path)
        print("\n✅✅✅ ¡Éxito! El proceso de automatización ha finalizado por completo.")
        print(f"El reporte final y limpio ha sido creado en: {processed_report_path}")
    except Exception as e:
        print(f"\n❌❌❌ Ocurrió un error crítico durante la ejecución del procesamiento de Excel: {e}")